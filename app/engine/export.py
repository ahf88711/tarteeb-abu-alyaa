"""Excel and PDF export for ranking results."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _rank_label(result: dict) -> str:
    if result.get("rank_display"):
        return str(result["rank_display"])
    return str(result.get("rank")) if result.get("rank") is not None else "—"


def export_ranking_text(results: list[dict], summary: dict) -> str:
    """Plain Arabic text suitable for WhatsApp / clipboard."""
    lines = [
        "📋 ترتيب أبو علياء",
        "══════════════════",
        f"مرتّب: {summary.get('ranked_successfully', '—')} · "
        f"تعادل: {summary.get('tied', 0)} · "
        f"غير محسوم: {summary.get('unresolved', 0)}",
        "",
    ]
    for r in results:
        if r.get("rank") is None:
            continue
        lines.append(
            f"{_rank_label(r)}. {r.get('original_name', '')} — "
            f"أحدث: {r.get('latest_date') or '—'} "
            f"({r.get('date_count', 0)} تاريخ) — {r.get('status', '')}"
        )
    lines.append("")
    lines.append("القاعدة: الأقدم في أحدث تاريخ يتقدّم؛ عند التعادل يُفحص التاريخ السابق.")
    return "\n".join(lines)


def export_audit_json(results: list[dict], summary: dict, messages: list = None) -> bytes:
    """Full machine-readable audit package."""
    import json
    from datetime import datetime

    payload = {
        "app": "ترتيب أبو علياء",
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "summary": summary,
        "results": results,
        "messages": messages or [],
        "rule": {
            "order": "newest_to_oldest_sequence",
            "compare": "older_wins_at_first_difference",
            "no_invented_tiebreakers": True,
        },
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def export_master_index(people: list[dict]) -> bytes:
    """Export master people index (name, dates, pages) as Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "فهرس الملف الرئيسي"
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "ترتيب أبو علياء — فهرس الملف الرئيسي"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1B4F72")
    ws.merge_cells("A1:E1")
    headers = ["م", "الاسم", "الرتبة", "الصفحات", "التواريخ (الأحدث ← الأقدم)"]
    header_fill = PatternFill("solid", fgColor="1B4F72")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", readingOrder=2)
    for i, p in enumerate(people, 1):
        ws.cell(3 + i, 1, i)
        ws.cell(3 + i, 2, p.get("name") or p.get("original_name"))
        ws.cell(3 + i, 3, p.get("rank_title") or "")
        pages = p.get("pages") or []
        ws.cell(3 + i, 4, "، ".join(str(x) for x in pages))
        dates = p.get("dates") or []
        ws.cell(3 + i, 5, " | ".join(dates))
    for col, w in enumerate([6, 28, 12, 14, 70], 1):
        ws.column_dimensions[chr(64 + col)].width = w
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_excel(results: list[dict], summary: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "الترتيب"
    ws.sheet_view.rightToLeft = True

    ws["A1"] = "ترتيب أبو علياء"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1B4F72")
    ws.merge_cells("A1:G1")

    ws["A2"] = (
        f"المطلوب: {summary.get('target_total', '—')} | "
        f"مؤكد: {summary.get('target_verified', '—')} | "
        f"مرتّب: {summary.get('ranked_successfully', '—')} | "
        f"تعادل: {summary.get('tied', '—')} | "
        f"غير محسوم: {summary.get('unresolved', '—')}"
    )
    ws.merge_cells("A2:G2")

    headers = [
        "الترتيب",
        "الاسم",
        "أحدث تاريخ",
        "التاريخ السابق",
        "عدد التواريخ",
        "الحالة",
        "سبب الترتيب",
    ]
    header_fill = PatternFill("solid", fgColor="1B4F72")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", readingOrder=2)

    body = Font(name="Arial", size=10)
    for i, r in enumerate(results):
        row = 5 + i
        vals = [
            _rank_label(r),
            r.get("original_name"),
            r.get("latest_date") or "—",
            r.get("previous_date") or "—",
            r.get("date_count") or 0,
            r.get("status"),
            r.get("explanation") or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row, col, v)
            cell.font = body
            cell.border = thin
            cell.alignment = Alignment(
                horizontal="center" if col < 7 else "right",
                wrap_text=True,
                readingOrder=2,
                vertical="center",
            )
        ws.row_dimensions[row].height = 40

    widths = [10, 28, 14, 14, 12, 22, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Dates sheet
    ws2 = wb.create_sheet("التواريخ التفصيلية")
    ws2.sheet_view.rightToLeft = True
    ws2["A1"] = "الاسم"
    ws2["B1"] = "التواريخ (الأحدث ← الأقدم)"
    for i, r in enumerate(results):
        ws2.cell(i + 2, 1, r.get("original_name"))
        dates = r.get("dates") or []
        ws2.cell(i + 2, 2, " | ".join(dates))
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 80

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _arabic_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = (
        Path.home() / ".fonts_ar" / "formal",
        Path("/usr/share/fonts/truetype/noto"),
        Path("/usr/share/fonts/opentype/noto"),
    )
    for font_dir in candidates:
        regular = font_dir / "NotoNaskhArabic-Regular.ttf"
        bold = font_dir / "NotoNaskhArabic-Bold.ttf"
        if regular.is_file():
            if "Ar" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("Ar", str(regular)))
            if "ArB" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(
                    TTFont("ArB", str(bold if bold.is_file() else regular))
                )
            return "Ar", "ArB"
    return "Helvetica", "Helvetica-Bold"


def _ar(text: str, fn: str) -> str:
    if fn == "Helvetica":
        return str(text)
    import arabic_reshaper
    from bidi.algorithm import get_display

    return get_display(arabic_reshaper.reshape(str(text)))


def export_pdf_simple(results: list[dict], summary: dict) -> bytes:
    """Arabic PDF via reportlab + reshaper."""
    from reportlab.lib.colors import HexColor, white
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    fn, fnb = _arabic_fonts()

    def P(text, style):
        return Paragraph(_ar(text, fn), style)

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )
    st_title = ParagraphStyle(
        "t", fontName=fnb, fontSize=14, alignment=TA_CENTER, textColor=HexColor("#1B4F72"), leading=20
    )
    st_h = ParagraphStyle("h", fontName=fnb, fontSize=8, alignment=TA_CENTER, textColor=white, leading=11)
    st_c = ParagraphStyle("c", fontName=fn, fontSize=7.5, alignment=TA_CENTER, leading=10)
    st_r = ParagraphStyle("r", fontName=fn, fontSize=7, alignment=TA_RIGHT, leading=10)

    page_w, _ = landscape(A4)
    usable = page_w - 2 * cm
    header = [
        P("سبب الترتيب", st_h),
        P("الحالة", st_h),
        P("العدد", st_h),
        P("السابق", st_h),
        P("الأحدث", st_h),
        P("الاسم", st_h),
        P("م", st_h),
    ]
    data = [header]
    for r in results:
        data.append(
            [
                P((r.get("explanation") or "")[:180], st_r),
                P(r.get("status") or "", st_c),
                P(str(r.get("date_count") or 0), st_c),
                P(r.get("previous_date") or "—", st_c),
                P(r.get("latest_date") or "—", st_c),
                P(r.get("original_name") or "", st_c),
                P(_rank_label(r), st_c),
            ]
        )

    col_w = [
        usable * 0.38,
        usable * 0.12,
        usable * 0.06,
        usable * 0.1,
        usable * 0.1,
        usable * 0.18,
        usable * 0.06,
    ]
    table = Table(data, colWidths=col_w, repeatRows=1)
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1B4F72")),
        ("GRID", (0, 0), (-1, -1), 0.3, HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for i in range(1, len(data)):
        if i % 2 == 0:
            cmds.append(("BACKGROUND", (0, i), (-1, i), HexColor("#EAF2F8")))
    table.setStyle(TableStyle(cmds))

    story = [
        P("ترتيب أبو علياء", st_title),
        Spacer(1, 3 * mm),
        P(
            f"مرتّب: {summary.get('ranked_successfully', '—')} | "
            f"تعادل: {summary.get('tied', '—')} | "
            f"غير محسوم: {summary.get('unresolved', '—')}",
            ParagraphStyle("s", fontName=fn, fontSize=9, alignment=TA_CENTER, leading=12),
        ),
        Spacer(1, 4 * mm),
        table,
    ]
    doc.build(story)
    return bio.getvalue()


def export_pdf_formal(
    results: list[dict],
    summary: dict,
    *,
    unit_title: str = "",
    prepared_by: str = "",
) -> bytes:
    """
    Formal portrait A4 ranking sheet with letterhead, stats, table,
    and signature lines — suitable for printing/official attachment.
    """
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    fn, fnb = _arabic_fonts()

    def P(text, style):
        return Paragraph(_ar(text, fn), style)

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        rightMargin=1.4 * cm,
        leftMargin=1.4 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.4 * cm,
    )
    page_w, _ = A4
    usable = page_w - 2.8 * cm

    st_org = ParagraphStyle(
        "org", fontName=fnb, fontSize=11, alignment=TA_CENTER, textColor=HexColor("#1B4F72"), leading=16
    )
    st_title = ParagraphStyle(
        "title", fontName=fnb, fontSize=16, alignment=TA_CENTER, textColor=HexColor("#0E3A56"), leading=22
    )
    st_sub = ParagraphStyle(
        "sub", fontName=fn, fontSize=9, alignment=TA_CENTER, textColor=HexColor("#566573"), leading=13
    )
    st_h = ParagraphStyle("h", fontName=fnb, fontSize=8, alignment=TA_CENTER, textColor=white, leading=11)
    st_c = ParagraphStyle("c", fontName=fn, fontSize=8, alignment=TA_CENTER, leading=11)
    st_note = ParagraphStyle(
        "note", fontName=fn, fontSize=8, alignment=TA_RIGHT, textColor=HexColor("#5D6D7E"), leading=12
    )

    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    ranked_only = [r for r in results if r.get("rank") is not None]

    header_cells = [
        P("الحالة", st_h),
        P("عدد التواريخ", st_h),
        P("التاريخ السابق", st_h),
        P("أحدث تاريخ", st_h),
        P("الاسم", st_h),
        P("الترتيب", st_h),
    ]
    data = [header_cells]
    for r in ranked_only:
        data.append(
            [
                P(r.get("status") or "", st_c),
                P(str(r.get("date_count") or 0), st_c),
                P(r.get("previous_date") or "—", st_c),
                P(r.get("latest_date") or "—", st_c),
                P(r.get("original_name") or "", st_c),
                P(_rank_label(r), st_c),
            ]
        )

    col_w = [
        usable * 0.16,
        usable * 0.12,
        usable * 0.16,
        usable * 0.16,
        usable * 0.28,
        usable * 0.12,
    ]
    table = Table(data, colWidths=col_w, repeatRows=1)
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1B4F72")),
        ("GRID", (0, 0), (-1, -1), 0.4, HexColor("#AAB7B8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for i in range(1, len(data)):
        if i % 2 == 0:
            cmds.append(("BACKGROUND", (0, i), (-1, i), HexColor("#EBF5FB")))
    table.setStyle(TableStyle(cmds))

    # Signature block
    st_sig = ParagraphStyle("sig", fontName=fn, fontSize=9, alignment=TA_CENTER, leading=14)
    sig = Table(
        [
            [
                P("التوقيع: ………………", st_sig),
                P("الاسم: ………………", st_sig),
                P("الرتبة: ………………", st_sig),
            ],
            [
                P("التاريخ: ………………", st_sig),
                P(prepared_by or "معدّ البيان: ………………", st_sig),
                P("الختم", st_sig),
            ],
        ],
        colWidths=[usable / 3.0] * 3,
    )
    sig.setStyle(
        TableStyle(
            [
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )

    story = [
        P(unit_title or "بيان ترتيب الأفراد حسب السجلات التاريخية", st_org),
        Spacer(1, 2 * mm),
        P("ترتيب أبو علياء", st_title),
        Spacer(1, 1 * mm),
        HRFlowable(width="100%", thickness=1.2, color=HexColor("#1B4F72"), spaceBefore=2, spaceAfter=6),
        P(
            f"تاريخ الإصدار: {today} · "
            f"عدد المرتّبين: {summary.get('ranked_successfully', len(ranked_only))} · "
            f"تعادل: {summary.get('tied', 0)} · "
            f"غير محسوم: {summary.get('unresolved', 0)}",
            st_sub,
        ),
        Spacer(1, 4 * mm),
        table,
        Spacer(1, 6 * mm),
        P(
            "قاعدة الترتيب: يُقارن أحدث تاريخ لكل فرد؛ عند التساوي يُنتقل للتاريخ السابق، "
            "والأقدم يتقدّم عند أول اختلاف. لا يُختلق كسر تعادل.",
            st_note,
        ),
        Spacer(1, 10 * mm),
        sig,
    ]
    doc.build(story)
    return bio.getvalue()
