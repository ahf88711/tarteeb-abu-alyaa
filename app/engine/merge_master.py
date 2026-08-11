"""Merge multiple master sources (PDF pages / Excel) into one people index."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable, Optional

from .extract_master import _merge_near_duplicate_people, extract_master_pdf
from .models import MasterPerson, make_master_key


def merge_people_dicts(
    *dicts: dict[str, MasterPerson],
) -> dict[str, MasterPerson]:
    """Union all master people, merging by soft identity."""
    combined: dict[str, MasterPerson] = {}
    for d in dicts:
        for key, person in d.items():
            if key not in combined:
                combined[key] = person
                continue
            # same normalized key — merge occurrences
            primary = combined[key]
            for notes in person.notes_texts:
                if notes not in primary.notes_texts:
                    primary.notes_texts.append(notes)
            for dt in person.dates:
                if not any(
                    x.normalized.iso() == dt.normalized.iso() for x in primary.dates
                ):
                    primary.dates.append(dt)
            for page in person.pages:
                if page not in primary.pages:
                    primary.pages.append(page)
            if person.rank_title and not primary.rank_title:
                primary.rank_title = person.rank_title
    return _merge_near_duplicate_people(combined)


def extract_master_pdfs(paths: Iterable[Path]) -> dict[str, MasterPerson]:
    """Process multiple PDFs (e.g. page scans) and merge."""
    acc: dict[str, MasterPerson] = {}
    page_offset = 0
    for path in paths:
        people = extract_master_pdf(path)
        # re-number pages to be unique across files
        for p in people.values():
            p.pages = [pg + page_offset for pg in p.pages]
            for d in p.dates:
                d.page = (d.page or 0) + page_offset
        max_page = 0
        for p in people.values():
            if p.pages:
                max_page = max(max_page, max(p.pages))
        page_offset = max(page_offset, max_page)
        acc = merge_people_dicts(acc, people)
    return acc


def extract_master_excel(path: Path) -> dict[str, MasterPerson]:
    """
    Import a structured master from Excel.
    Expected columns (Arabic or English headers):
      الاسم / name | الملاحظات / notes | الصفحة / page | الرتبة / rank
    Or: الاسم | تواريخ (pipe/comma separated)
    """
    from openpyxl import load_workbook

    from .dates import extract_all_dates, parse_hijri_date, ExtractedDate

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(c or "").strip() for c in rows[0]]
    header_l = [h.lower() for h in header]

    def col(*names: str) -> Optional[int]:
        for i, h in enumerate(header):
            hl = h.lower()
            for n in names:
                if n in h or n in hl:
                    return i
        return None

    i_name = col("الاسم", "اسم", "name")
    i_notes = col("الملاحظات", "ملاحظات", "notes")
    i_dates = col("التواريخ", "تواريخ", "dates")
    i_page = col("الصفحة", "صفحات", "page")
    i_rank = col("الرتبة", "رتبة", "rank")

    if i_name is None:
        # try first column
        i_name = 0

    people: dict[str, MasterPerson] = {}
    for row in rows[1:]:
        if not row or i_name is None:
            continue
        raw_name = row[i_name]
        if raw_name is None or not str(raw_name).strip():
            continue
        name = str(raw_name).strip()
        key = make_master_key(name)
        if key not in people:
            people[key] = MasterPerson(original_name=name, normalized_name=key)
        person = people[key]

        rank = ""
        if i_rank is not None and row[i_rank]:
            rank = str(row[i_rank]).strip()
            if not person.rank_title:
                person.rank_title = rank

        page = 0
        if i_page is not None and row[i_page] is not None:
            try:
                page = int(row[i_page])
            except (TypeError, ValueError):
                page = 0

        notes = ""
        if i_notes is not None and row[i_notes]:
            notes = str(row[i_notes])
        if notes:
            person.add_occurrence(
                page=page or 1,
                notes=notes,
                rank_title=rank,
                date_confidence=0.98,
            )

        if i_dates is not None and row[i_dates]:
            raw = str(row[i_dates])
            parts = [p.strip() for p in re.split(r"[|,;]+", raw) if p.strip()]
            for part in parts:
                hd = parse_hijri_date(part)
                if not hd:
                    # try extract_all
                    for d in extract_all_dates(part, page=page or 1, confidence=0.98, person_name=name):
                        if not any(x.normalized.iso() == d.normalized.iso() for x in person.dates):
                            d.verified = True
                            person.dates.append(d)
                    continue
                if not any(x.normalized.iso() == hd.iso() for x in person.dates):
                    person.dates.append(
                        ExtractedDate(
                            normalized=hd,
                            original_text=part,
                            page=page or 1,
                            confidence=1.0,
                            verified=True,
                            source_snippet="Excel",
                            person_name=name,
                        )
                    )
                if page and page not in person.pages:
                    person.pages.append(page)

    return _merge_near_duplicate_people(people)


def rename_master_person(
    people: dict[str, MasterPerson],
    old_key: str,
    new_name: str,
) -> dict[str, MasterPerson]:
    """User correction of a master identity display/name."""
    new_name = (new_name or "").strip()
    if not new_name:
        return people
    # find by key or soft match
    person = people.get(old_key)
    if not person:
        for k, p in people.items():
            if p.original_name == old_key or p.normalized_name == make_master_key(old_key):
                person = p
                old_key = k
                break
    if not person:
        return people

    new_key = make_master_key(new_name)
    del people[old_key]
    person.original_name = new_name
    person.normalized_name = new_key
    if new_key in people:
        # merge into existing
        people = merge_people_dicts(people, {new_key: person})
    else:
        people[new_key] = person
    return people
