"""End-to-end pipeline orchestration."""

from __future__ import annotations

import uuid
from pathlib import Path
from typing import Optional

from .dates import unique_dates_newest_first
from .extract_master import extract_master_pdf
from .extract_targets import extract_target_names
from .merge_master import (
    extract_master_excel,
    extract_master_pdfs,
    merge_people_dicts,
    rename_master_person,
)
from .models import NameStatus, SessionState, make_master_key
from .ranking import RankPerson, rank_people, summarize_results


def new_session() -> SessionState:
    return SessionState(session_id=str(uuid.uuid4())[:12], phase="init")


def set_progress(session: SessionState, phase: str, message: str, pct: int = 0) -> None:
    session.phase = phase
    session.messages.append(message)
    session.summary = {
        **session.summary,
        "progress_pct": max(0, min(100, pct)),
        "progress_message": message,
    }


def load_master(session: SessionState, pdf_path: Path) -> SessionState:
    session.master_path = str(pdf_path)
    set_progress(session, "master_loading", "جاري معالجة الملف الرئيسي بالكامل…", 10)
    people = extract_master_pdf(pdf_path)
    session.master_people = people
    set_progress(
        session,
        "master_loaded",
        f"تم استخراج {len(people)} شخصًا من الملف الرئيسي عبر جميع الصفحات.",
        100,
    )
    return session


def load_master_many(session: SessionState, paths: list[Path]) -> SessionState:
    """Load one or more master files (PDF/Excel) and merge."""
    if not paths:
        set_progress(session, "error", "لم يُرفع أي ملف رئيسي.", 0)
        return session

    set_progress(session, "master_loading", f"جاري معالجة {len(paths)} ملف(ات) رئيسية…", 5)
    pdfs: list[Path] = []
    excels: list[Path] = []
    for p in paths:
        suf = p.suffix.lower()
        if suf == ".pdf":
            pdfs.append(p)
        elif suf in {".xlsx", ".xlsm", ".xls"}:
            excels.append(p)
        else:
            session.messages.append(f"تم تجاهل صيغة غير مدعومة: {p.name}")

    people: dict = {}
    total = max(1, len(pdfs) + len(excels))
    done = 0

    if pdfs:
        set_progress(session, "master_loading", f"استخراج PDF ({len(pdfs)})…", 15)
        people = merge_people_dicts(people, extract_master_pdfs(pdfs))
        done += len(pdfs)
        set_progress(
            session,
            "master_loading",
            f"أُنجز {done}/{total} — الأشخاص حتى الآن: {len(people)}",
            int(15 + 70 * done / total),
        )

    for xp in excels:
        set_progress(session, "master_loading", f"استيراد Excel: {xp.name}", int(15 + 70 * done / total))
        people = merge_people_dicts(people, extract_master_excel(xp))
        done += 1

    session.master_people = people
    session.master_path = str(paths[0]) if len(paths) == 1 else f"{len(paths)} files"
    set_progress(
        session,
        "master_loaded",
        f"تم دمج {len(paths)} ملف — الفهرس النهائي: {len(people)} شخصًا.",
        100,
    )
    return session


def rename_person(session: SessionState, old_key: str, new_name: str) -> SessionState:
    session.master_people = rename_master_person(
        session.master_people, old_key, new_name
    )
    session.messages.append(f"تم تصحيح اسم في الملف الرئيسي إلى: {new_name}")
    return session


def _summarize_targets(session: SessionState) -> None:
    targets = session.target_names
    verified = sum(1 for t in targets if t.status == NameStatus.VERIFIED)
    review = sum(
        1
        for t in targets
        if t.status in (NameStatus.NEEDS_REVIEW, NameStatus.AMBIGUOUS)
    )
    unknown = sum(
        1
        for t in targets
        if t.status in (NameStatus.UNKNOWN, NameStatus.NOT_IN_MASTER)
    )
    session.summary = {
        **session.summary,
        "target_total": len(targets),
        "target_verified": verified,
        "target_needs_review": review,
        "target_not_in_master": unknown,
    }
    session.messages.append(
        f"أسماء مطلوبة: {len(targets)} — مؤكد: {verified}، مراجعة: {review}، غير موجود: {unknown}."
    )


def _append_targets(session: SessionState, new_targets: list) -> int:
    """Merge new targets into session without duplicating normalized names."""
    existing = {t.normalized_name for t in session.target_names}
    added = 0
    for t in new_targets:
        if t.normalized_name in existing:
            continue
        existing.add(t.normalized_name)
        session.target_names.append(t)
        added += 1
    return added


def load_targets(session: SessionState, path: Path, *, replace: bool = True) -> SessionState:
    if not session.master_people:
        session.messages.append("يجب رفع الملف الرئيسي أولًا لتدقيق الأسماء.")
        session.phase = "error"
        return session
    session.target_path = str(path)
    session.messages.append("جاري استخراج قائمة الأسماء المطلوبة مع التحقق…")
    suf = path.suffix.lower()
    if suf in {".xlsx", ".xlsm"}:
        targets = _extract_targets_excel(path, session.master_people)
    else:
        targets = extract_target_names(path, session.master_people)
    if replace:
        session.target_names = targets
        added = len(targets)
    else:
        added = _append_targets(session, targets)
    session.phase = "names_extracted"
    session.messages.append(f"أُضيف/حُدّث {added} اسمًا من {path.name}.")
    _summarize_targets(session)
    return session


def load_targets_many(session: SessionState, paths: list[Path]) -> SessionState:
    if not session.master_people:
        session.messages.append("يجب رفع الملف الرئيسي أولًا.")
        session.phase = "error"
        return session
    first = True
    for path in paths:
        load_targets(session, path, replace=first)
        first = False
    return session


def _extract_targets_excel(path: Path, master: dict) -> list:
    """Read target names from Excel column الاسم / name."""
    from openpyxl import load_workbook
    from .extract_targets import match_to_master
    from .models import TargetName, NameStatus
    import uuid as _uuid

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    name_col = 0
    for i, h in enumerate(header):
        if any(x in h for x in ("الاسم", "اسم", "name", "Name")):
            name_col = i
            break
    start = 1 if any(x in header[name_col] for x in ("الاسم", "اسم", "name", "Name")) else 0
    results = []
    seen = set()
    for row in rows[start:]:
        if not row or name_col >= len(row) or row[name_col] is None:
            continue
        name = str(row[name_col]).strip()
        if len(name) < 3:
            continue
        key = make_master_key(name)
        if key in seen:
            continue
        seen.add(key)
        status, conf, candidates, matched = match_to_master(name, master)
        if key in master:
            status = NameStatus.VERIFIED
            conf = 1.0
            matched = master[key].original_name
        results.append(
            TargetName(
                id=str(_uuid.uuid4())[:8],
                original_name=name,
                normalized_name=key,
                ocr_raw=name,
                confidence=round(conf, 4),
                status=status,
                candidates=candidates,
                matched_master_name=matched,
            )
        )
    return results


def add_manual_target_names(session: SessionState, names: list[str]) -> SessionState:
    """
    Add target names typed/pasted by the user (one per line).
    Still matched against master — never ranks unknown people as verified without match.
    """
    from .extract_targets import match_to_master
    from .models import TargetName
    from .normalize import normalize_arabic_name

    if not session.master_people:
        session.messages.append("يجب رفع الملف الرئيسي أولًا.")
        session.phase = "error"
        return session

    existing = {t.normalized_name for t in session.target_names}
    added = 0
    for raw in names:
        name = (raw or "").strip()
        if not name or len(name) < 3:
            continue
        key = make_master_key(name)
        if key in existing:
            continue
        existing.add(key)
        status, conf, candidates, matched = match_to_master(name, session.master_people)
        if key in session.master_people:
            status = NameStatus.VERIFIED
            conf = 1.0
            matched = session.master_people[key].original_name
        session.target_names.append(
            TargetName(
                id=str(uuid.uuid4())[:8],
                original_name=name,
                normalized_name=key,
                ocr_raw=name,
                confidence=round(conf, 4),
                status=status,
                candidates=candidates,
                matched_master_name=matched,
            )
        )
        added += 1

    session.phase = "names_extracted"
    session.messages.append(f"أُضيف {added} اسمًا يدويًا.")
    _summarize_targets(session)
    return session


def auto_confirm_high_confidence(
    session: SessionState, *, min_confidence: float = 0.92
) -> SessionState:
    """
    Confirm only high-confidence / exact master matches.
    Never confirms ambiguous or multi-candidate close races.
    """
    from .models import NameStatus

    corrections = []
    for t in session.target_names:
        if t.status == NameStatus.AMBIGUOUS:
            continue
        if t.status == NameStatus.VERIFIED and t.matched_master_name:
            corrections.append({"id": t.id, "action": "confirm"})
            continue
        if (
            t.matched_master_name
            and t.confidence >= min_confidence
            and t.status != NameStatus.NOT_IN_MASTER
        ):
            # require clear winner: top candidate gap
            cands = t.candidates or []
            if len(cands) >= 2:
                gap = float(cands[0].get("confidence", 0)) - float(
                    cands[1].get("confidence", 0)
                )
                if gap < 0.08:
                    continue  # too close — human review
            corrections.append(
                {"id": t.id, "action": "set_name", "name": t.matched_master_name}
            )
    if corrections:
        apply_name_corrections(session, corrections)
        session.messages.append(
            f"تأكيد تلقائي حذر لـ {len(corrections)} اسمًا (ثقة ≥ {min_confidence:.0%} وفجوة واضحة)."
        )
    else:
        session.messages.append("لا أسماء مؤهلة للتأكيد التلقائي الحذر.")
    return session


def apply_name_corrections(
    session: SessionState, corrections: list[dict]
) -> SessionState:
    """
    corrections: [{id, action: confirm|set_name|choose_candidate, name?: str}]
    """
    by_id = {t.id: t for t in session.target_names}
    for c in corrections:
        t = by_id.get(c.get("id"))
        if not t:
            continue
        action = c.get("action")
        if action == "confirm" and t.matched_master_name:
            t.status = NameStatus.VERIFIED
            t.user_corrected_name = t.matched_master_name
        elif action == "set_name":
            name = (c.get("name") or "").strip()
            if name:
                t.user_corrected_name = name
                t.normalized_name = make_master_key(name)
                key = t.normalized_name
                if key in session.master_people:
                    t.status = NameStatus.VERIFIED
                    t.matched_master_name = session.master_people[key].original_name
                else:
                    # fuzzy re-check
                    from .extract_targets import match_to_master

                    st, conf, cands, matched = match_to_master(name, session.master_people)
                    t.status = st
                    t.confidence = conf
                    t.candidates = cands
                    t.matched_master_name = matched
        elif action == "choose_candidate":
            name = (c.get("name") or "").strip()
            if name:
                t.user_corrected_name = name
                t.normalized_name = make_master_key(name)
                key = t.normalized_name
                if key in session.master_people:
                    t.matched_master_name = session.master_people[key].original_name
                    t.status = NameStatus.VERIFIED
                else:
                    # Never mark verified if the chosen identity is absent from master
                    from .extract_targets import match_to_master

                    st, conf, cands, matched = match_to_master(name, session.master_people)
                    t.status = st
                    t.confidence = conf
                    t.candidates = cands
                    t.matched_master_name = matched
                    if st != NameStatus.VERIFIED:
                        # User explicitly chose a label — still require master presence
                        if matched and make_master_key(matched) in session.master_people:
                            t.matched_master_name = matched
                            t.normalized_name = make_master_key(matched)
                            t.status = NameStatus.VERIFIED
                        else:
                            t.status = NameStatus.NOT_IN_MASTER
        elif action == "reject":
            t.status = NameStatus.UNKNOWN
    session.phase = "names_reviewed"
    session.messages.append("تم تحديث مراجعة الأسماء.")
    _summarize_targets(session)
    return session


def collect_dates_for_targets(session: SessionState) -> SessionState:
    """Attach master dates to verified target names only."""
    for t in session.target_names:
        if t.status != NameStatus.VERIFIED:
            continue
        key = make_master_key(t.display_name)
        person = session.master_people.get(key)
        if not person and t.matched_master_name:
            person = session.master_people.get(make_master_key(t.matched_master_name))
        if not person:
            t.status = NameStatus.NOT_IN_MASTER
    session.phase = "dates_ready"
    session.messages.append("تم جمع التواريخ من الملف الرئيسي للأسماء المؤكدة.")
    return session


def bulk_verify_safe_dates(session: SessionState) -> SessionState:
    """Verify all high-confidence dates that are not flagged for review."""
    n = 0
    for person in session.master_people.values():
        for d in person.dates:
            if not d.needs_review and d.confidence >= 0.85:
                if not d.verified:
                    d.verified = True
                    n += 1
    session.messages.append(f"اعتُمد {n} تاريخًا عالي الثقة تلقائيًا (بدون المشبوهة).")
    return session


def apply_date_reviews(session: SessionState, reviews: list[dict]) -> SessionState:
    """
    reviews: [{master_key, action: verify_all|verify_date|delete_date|add_date,
               date?: str, original_text?: str}]
    """
    from .dates import parse_hijri_date, ExtractedDate

    for r in reviews:
        key = make_master_key(r.get("master_key") or r.get("name") or "")
        person = session.master_people.get(key)
        if not person:
            # soft find
            from .normalize import soft_normalize_for_fuzzy

            soft = soft_normalize_for_fuzzy(key)
            for mk, mp in session.master_people.items():
                if soft_normalize_for_fuzzy(mp.original_name) == soft:
                    person = mp
                    break
        if not person:
            continue
        action = r.get("action")
        if action == "verify_all":
            for d in person.dates:
                d.verified = True
                d.needs_review = False
        elif action == "verify_date":
            iso = r.get("date")
            for d in person.dates:
                if d.normalized.iso() == iso or d.normalized.display() == iso:
                    d.verified = True
                    d.needs_review = False
        elif action == "delete_date":
            iso = r.get("date")
            person.dates = [
                d
                for d in person.dates
                if d.normalized.iso() != iso and d.normalized.display() != iso
            ]
        elif action == "add_date":
            raw = r.get("date") or r.get("original_text") or ""
            hd = parse_hijri_date(raw)
            if hd:
                person.dates.append(
                    ExtractedDate(
                        normalized=hd,
                        original_text=raw,
                        page=int(r.get("page") or 0),
                        confidence=1.0,
                        verified=True,
                        source_snippet="أضافه المستخدم",
                        person_name=person.original_name,
                    )
                )
    session.phase = "dates_reviewed"
    session.messages.append("تم تحديث مراجعة التواريخ.")
    return session


def run_ranking(session: SessionState, *, auto_verify_dates: bool = True) -> SessionState:
    """
    Deterministic ranking of VERIFIED target names only.
    """
    rank_people_list: list[RankPerson] = []
    skipped_review = 0
    not_found = 0
    seen_master_keys: set[str] = set()

    for t in session.target_names:
        if t.status in (NameStatus.NEEDS_REVIEW, NameStatus.AMBIGUOUS):
            skipped_review += 1
            continue
        if t.status in (NameStatus.NOT_IN_MASTER, NameStatus.UNKNOWN):
            not_found += 1
            continue
        if t.status != NameStatus.VERIFIED:
            skipped_review += 1
            continue

        key = make_master_key(t.display_name)
        master = session.master_people.get(key)
        if not master and t.matched_master_name:
            master = session.master_people.get(make_master_key(t.matched_master_name))

        if not master:
            # Soft lookup ONLY if exactly one master identity shares the soft key.
            # Multiple soft collisions → refuse (false match worse than no match).
            from .normalize import soft_normalize_for_fuzzy

            soft = soft_normalize_for_fuzzy(t.display_name)
            soft_hits = [
                (mk, mp)
                for mk, mp in session.master_people.items()
                if soft_normalize_for_fuzzy(mp.original_name) == soft
            ]
            if len(soft_hits) == 1:
                key, master = soft_hits[0]
            elif len(soft_hits) > 1:
                # Ambiguous soft identity — do not rank
                skipped_review += 1
                t.status = NameStatus.AMBIGUOUS
                continue

        if not master:
            not_found += 1
            continue

        # One ranking row per master person (avoid duplicate OCR targets)
        if master.normalized_name in seen_master_keys:
            continue
        seen_master_keys.add(master.normalized_name)

        if auto_verify_dates:
            for d in master.dates:
                # Never auto-verify dates flagged for review / low confidence
                if d.needs_review:
                    d.verified = False
                elif d.confidence >= 0.85:
                    d.verified = True

        dates = unique_dates_newest_first(master.dates, only_verified=True)
        pending_review = sum(1 for d in master.dates if d.needs_review)
        rank_people_list.append(
            RankPerson(
                id=t.id,
                original_name=master.original_name,
                normalized_name=master.normalized_name,
                dates=dates,
                meta={
                    "target_original": t.original_name,
                    "pages": master.pages,
                    "all_dates": [d.to_dict() for d in master.dates],
                    "notes": master.notes_texts,
                    "pending_date_review_count": pending_review,
                    "verified_date_count": len(dates),
                },
            )
        )

    entries = rank_people(rank_people_list)
    session.ranking_results = []
    pending_total = 0
    for e in entries:
        d = e.to_dict()
        d["meta"] = e.person.meta
        pending_total += int((e.person.meta or {}).get("pending_date_review_count") or 0)
        # Surface warning when ranked on thin/uncertain date evidence
        if e.person.dates and (e.person.meta or {}).get("pending_date_review_count", 0) > 0:
            d["explanation"] = (
                d.get("explanation") or ""
            ) + f" ⚠ يوجد {(e.person.meta or {}).get('pending_date_review_count')} تاريخًا بحاجة مراجعة."
        session.ranking_results.append(d)
    if pending_total:
        session.messages.append(
            f"تنبيه: {pending_total} تاريخًا ما زال بحاجة مراجعة بشرية — راجع قبل الاعتماد النهائي."
        )
    session.summary["pending_date_review_total"] = pending_total

    # Append non-ranked targets for transparency
    for t in session.target_names:
        if any(r["id"] == t.id for r in session.ranking_results):
            continue
        status = t.status.value
        session.ranking_results.append(
            {
                "rank": None,
                "id": t.id,
                "original_name": t.display_name,
                "normalized_name": t.normalized_name,
                "status": status,
                "explanation": "لم يُدرَج في الترتيب بسبب حالة الاسم أو غياب التواريخ.",
                "dates": [],
                "date_count": 0,
                "latest_date": None,
                "previous_date": None,
            }
        )

    stats = summarize_results(entries)
    session.summary = {
        **session.summary,
        **stats,
        "skipped_needs_review": skipped_review,
        "not_found": not_found,
        "ranked_candidates": len(rank_people_list),
    }
    session.phase = "ranked"
    session.messages.append(
        f"اكتمل الترتيب: {stats['ranked_successfully']} مرتّب، "
        f"{stats['tied']} تعادل، {stats['unresolved']} غير محسوم، "
        f"{stats['no_dates']} بدون تواريخ."
    )
    # Optional auto-export to Desktop for convenience (timestamped + latest)
    try:
        from datetime import datetime
        from pathlib import Path
        from .export import (
            export_audit_json,
            export_excel,
            export_pdf_formal,
            export_ranking_text,
        )

        desk = Path.home() / "Desktop" / "ترتيب_أبو_علياء"
        desk.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx = export_excel(session.ranking_results, session.summary)
        pdf = export_pdf_formal(session.ranking_results, session.summary)
        txt = export_ranking_text(session.ranking_results, session.summary)
        audit = export_audit_json(
            session.ranking_results, session.summary, session.messages[-30:]
        )
        (desk / "الترتيب.xlsx").write_bytes(xlsx)
        (desk / "الترتيب_رسمي.pdf").write_bytes(pdf)
        (desk / "الترتيب.txt").write_text(txt, encoding="utf-8")
        (desk / "تدقيق.json").write_bytes(audit)
        # keep history copies
        hist = desk / "سجل"
        hist.mkdir(exist_ok=True)
        (hist / f"ترتيب_{stamp}.xlsx").write_bytes(xlsx)
        (hist / f"ترتيب_{stamp}.pdf").write_bytes(pdf)
        (hist / f"ترتيب_{stamp}.txt").write_text(txt, encoding="utf-8")
        (hist / f"تدقيق_{stamp}.json").write_bytes(audit)
        session.messages.append(f"نُسخت النتائج تلقائيًا إلى: {desk}")
        session.summary["auto_export_dir"] = str(desk)
    except Exception as e:
        session.messages.append(f"تعذّر النسخ التلقائي لسطح المكتب: {e}")
    return session


def person_audit(session: SessionState, person_id: str) -> Optional[dict]:
    for r in session.ranking_results:
        if r.get("id") == person_id:
            return r
    return None
